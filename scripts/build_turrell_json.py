import json
import re
import time
import urllib.parse
import urllib.request
from pathlib import Path
from urllib.error import HTTPError

from openpyxl import load_workbook


INPUT_XLSX = Path("/Users/kohei/Downloads/turrell_works_v3.xlsx")
OUTPUT_JSON = Path("/Users/kohei/msdk8734.github.io/data/turrell.json")

HEADERS = {
    "User-Agent": "Codex/1.0 (turrell map research)",
    "Accept": "application/json",
}

MANUAL_COORDS = {
    "Panza Collection": (45.8250204, 8.8292766, "high", "Villa Panza, Piazza Litta, Varese, Lombardia, Italia"),
    "Roden Crater (Skystone Foundation)": (35.4258751, -111.2588365, "high", "Roden Crater, Coconino County, Arizona, United States"),
    "Mattress Factory": (40.4569286, -80.0122738, "high", "Mattress Factory, 500 Sampsonia Street, Pittsburgh, Pennsylvania, United States"),
    "MoMA PS1": (40.7455823, -73.9474668, "high", "MoMA PS1, Long Island City, Queens, New York, United States"),
    "Liss Ard Estate": (51.5478232, -9.2662109, "medium", "Skibbereen, County Cork, Munster, Ireland"),
    "Israel Museum": (31.7718880, 35.2033867, "high", "Israel Museum, Jerusalem, Israel"),
    "Kijkduin (Stroom Den Haag)": (52.0631375, 4.2183879, "high", "Hemels Gewelf, Kijkduin, Den Haag, Netherlands"),
    "House of Light": (37.1892134, 138.7297989, "high", "House of Light, Tokamachi, Niigata, Japan"),
    "M HKA (Museum van Hedendaagse Kunst)": (51.2114872, 4.3898602, "high", "M HKA - Museum van Hedendaagse Kunst Antwerpen, Antwerp, Belgium"),
    "Cheekwood Botanical Garden & Museum of Art": (36.0855412, -86.8725270, "medium", "Cheekwood Estate & Gardens, Nashville, Tennessee, United States"),
    "Tremenheere Sculpture Gardens": (50.1384598, -5.5082894, "medium", "Tremenheere Sculpture Gardens, Cornwall, United Kingdom"),
    "Museum of Fine Arts Houston": (29.7251296, -95.3894511, "high", "Museum of Fine Arts, Houston, Texas, United States"),
    "Art Tower Mito": (36.3804847, 140.4653089, "high", "水戸芸術館, 水戸芸術館通り, 水戸市, 茨城県, 日本"),
    "Kielder Water": (55.1894281, -2.5279299, "high", "Kielder Water, Northumberland, United Kingdom"),
    "Live Oak Friends Meeting House": (29.8094585, -95.4233068, "high", "Live Oak Friends Meeting, West 26th Street, Houston, Texas, United States"),
    "Scottsdale Museum of Contemporary Art": (33.4949327, -111.9261427, "high", "Scottsdale Museum of Contemporary Art, Arizona, United States"),
    "James Turrell Museum, Bodega Colomé": (-25.5128416, -66.3922333, "medium", "Colomé Winery, Ruta Provincial 53, Molinos, Salta, Argentina"),
    "Contemporary Art Museum Kumamoto (熊本市現代美術館)": (32.8027895, 130.7081193, "high", "Contemporary Art Museum Kumamoto, Kumamoto, Japan"),
    "Henry Art Gallery": (47.6606073, -122.3096577, "high", "Henry Art Gallery, Seattle, Washington, United States"),
    "Sheats-Goldstein Residence": (34.0935898, -118.4351349, "medium", "Sheats-Goldstein Residence, Angelo View Drive, Los Angeles County, California, United States"),
    "21st Century Museum of Contemporary Art (金沢21世紀美術館)": (36.5609736, 136.6566384, "high", "21st Century Museum of Contemporary Art, Kanazawa, Japan"),
    "Chichu Art Museum (地中美術館)": (34.4498970, 133.9859026, "high", "Chichu Art Museum, Naoshima, Kagawa, Japan"),
    "de Young Museum": (37.7714781, -122.4686779, "high", "de Young Museum, San Francisco, California, United States"),
    "University of Illinois at Chicago": (41.8719492, -87.6492662, "high", "University of Illinois Chicago, Illinois, United States"),
    "Hotel Castell": (46.6032019, 9.9589546, "high", "Hotel Castell, Zuoz, Switzerland"),
    "Centre for International Light Art (CILA)": (51.5347899, 7.6898700, "high", "Zentrum für Internationale Lichtkunst, Unna, Germany"),
    "Fundación NMAC": (36.2765509, -5.9306216, "high", "Fundación NMAC, Cádiz, Spain"),
    "Walker Art Center": (44.9685097, -93.2885354, "high", "Walker Art Center, Minneapolis, Minnesota, United States"),
    "Museum der Moderne (Salzburg Foundation)": (47.7965359, 13.0388076, "high", "Museum der Moderne Salzburg, Austria"),
    "Nasher Sculpture Center": (32.7877682, -96.8002693, "high", "Nasher Sculpture Center, Dallas, Texas, United States"),
    "Stonescape Vineyard": (38.5634644, -122.5533869, "medium", "4301 Azalea Springs Way, Calistoga, California, United States"),
    "Yorkshire Sculpture Park": (53.6440184, -1.5799510, "high", "Yorkshire Sculpture Park, Wakefield, United Kingdom"),
    "Pomona College": (34.0973295, -117.7107296, "high", "Pomona College, Claremont, California, United States"),
    "National Gallery of Australia": (-35.3141458, 149.1489158, "high", "National Gallery of Australia, Canberra, Australia"),
    "Kulturcentrum (Kulturforum Järna)": (59.0683762, 17.6183139, "high", "Kulturcentrum Järna, Södertälje, Sweden"),
    "Ringling Museum of Art": (27.3816096, -82.5745213, "high", "The Ringling, Sarasota, Florida, United States"),
    "Crystal Bridges Museum of American Art": (36.3783715, -94.2022804, "high", "Crystal Bridges Museum of American Art, Arkansas, United States"),
    "Rice University": (29.7173945, -95.4018312, "high", "Rice University, Houston, Texas, United States"),
    "Arizona State University": (33.4213174, -111.9343659, "high", "Arizona State University, Tempe, Arizona, United States"),
    "Museum SAN": (37.4103991, 127.8194561, "medium", "260 Oak Valley 2-gil, Wonju-si, Gangwon-do, South Korea"),
    "Temple Hotel": (39.8883675, 116.4329098, "high", "Xi Zhao Si Temple Hotel, Beijing, China"),
    "Louis Vuitton, The Shops at Crystals": (36.1077414, -115.1742962, "high", "The Shops at Crystals, South Las Vegas Boulevard, Las Vegas, Nevada, United States"),
    "University of Texas at Austin": (30.2849185, -97.7340567, "high", "The University of Texas at Austin, Texas, United States"),
    "Ekebergparken Sculpture Park": (59.8983657, 10.7793902, "high", "Ekebergparken, Oslo, Norway"),
    "Kayne Griffin Corcoran Gallery": (34.0345372, -118.3500220, "medium", "1201 South La Brea Avenue, Los Angeles, California, United States"),
    "Chestnut Hill Friends Meeting House": (40.0683907, -75.1960659, "high", "Chestnut Hill Friends Meeting, East Mermaid Lane, Philadelphia, Pennsylvania, United States"),
    "Museum of Old and New Art (MONA)": (-42.8084777, 147.2507375, "high", "Museum of Old and New Art, Hobart, Australia"),
    "Jardín Botánico Culiacán": (24.8041880, -107.3897212, "high", "Jardín Botánico Culiacán, Mexico"),
    "Houghton Hall": (52.8257481, 0.6439119, "high", "Houghton Hall, Norfolk, United Kingdom"),
    "Dorotheenstädtischer Friedhof": (52.5278070, 13.3821319, "high", "Dorotheenstädtischer Friedhof, Berlin, Germany"),
    "Venet Foundation": (43.4707322, 6.5471689, "high", "Fondation Venet, Le Muy, France"),
    "Museum Voorlinden": (52.1204909, 4.3518012, "high", "Museum Voorlinden, Wassenaar, Netherlands"),
    "Oberlech": (47.2114751, 10.1644269, "high", "Skyspace Lech, Oberlech, Austria"),
    "Posada Ayana": (-34.8390710, -54.6414790, "medium", "Paseo del Marinero, José Ignacio, Uruguay"),
    "MASS MoCA": (42.7007012, -73.1084714, "high", "MASS MoCA, North Adams, Massachusetts, United States"),
    "Tec de Monterrey": (25.6516420, -100.2899956, "high", "Tecnológico de Monterrey, Nuevo León, Mexico"),
    "Green Box Arts, Red Butte": (38.9363281, -105.0137372, "medium", "Red Butte Recreation Area trailheads, Green Mountain Falls, Colorado, United States"),
    "Friends Seminary": (40.7342206, -73.9917160, "high", "Friends Seminary, Manhattan, New York, United States"),
    "Keith House": (32.7464972, -97.3307658, "high", "Fort Worth, Texas, United States"),
    "ARoS Aarhus Art Museum": (56.1538877, 10.1996046, "high", "ARoS Kunstmuseum, Vester Allé, Aarhus, Denmark"),
    "505 Fifth Avenue": (40.7538318, -73.9823387, "high", "505 Fifth Avenue, Manhattan, New York, United States"),
    "Caisse des Dépôts": (48.8559449, 2.3606973, "high", "Caisse des Dépôts, Paris, France"),
    "Cenote Santa Maria": (20.2465800, -88.1756500, "medium", "Tixcacalteyub, Yucatán, Mexico"),
    "Craiganour Estate": (56.7040500, -4.2621900, "medium", "PH17 2QN, Scotland, United Kingdom"),
    "Kilfane Gardens": (52.5493151, -7.0896208, "high", "Kilfane Glen & Waterfall, Thomastown, County Kilkenny, Ireland"),
    "De Pont Museum": (51.5477508, 5.0818711, "high", "De Pont Museum, Tilburg, Netherlands"),
    "Dornier Museum": (47.6809778, 9.5110167, "high", "Dornier Museum, Friedrichshafen, Germany"),
    "Dresdner Bank (now Commerzbank)": (50.1106444, 8.6820917, "medium", "Frankfurt, Germany"),
    "FIFA HQ": (47.3738780, 8.5416940, "high", "FIFA Headquarters, Zurich, Switzerland"),
    "Fontainebleau Miami Beach": (25.8176945, -80.1226982, "high", "Fontainebleau Miami Beach, Florida, United States"),
    "Glenstone Museum": (39.0166340, -77.2081759, "high", "Glenstone Museum, Potomac, Maryland, United States"),
    "Greenwich Academy": (41.0262654, -73.6284630, "high", "Greenwich Academy, Connecticut, United States"),
    "Hacienda Ochil": (20.7726580, -89.7059210, "medium", "Hacienda Ochil, Yucatán, Mexico"),
    "Kunsthalle Mannheim": (49.4825317, 8.4763610, "high", "Kunsthalle Mannheim, Germany"),
    "Kunsthalle Bremen": (53.0726239, 8.8192232, "high", "Kunsthalle Bremen, Germany"),
    "L'hoist Foundation": (50.6815750, 4.5823000, "high", "Lhoist Group, Limelette, Belgium"),
    "Le Mondrian Hotel": (34.0943171, -118.3768917, "high", "Mondrian Los Angeles, California, United States"),
    "MAK (Museum of Applied Arts)": (48.2067786, 16.3800348, "high", "MAK Museum of Applied Arts, Vienna, Austria"),
    "Sprengel Museum": (52.3611946, 9.7399001, "high", "Sprengel Museum Hannover, Germany"),
    "Münchener Rück (Munich RE)": (48.1469864, 11.6168402, "high", "Munich Re, Munich, Germany"),
    "MMK": (50.1100421, 8.6909159, "high", "Museum für Moderne Kunst, Frankfurt, Germany"),
    "Bay Adelaide Centre": (43.6487007, -79.3819882, "high", "Bay Adelaide Centre, Toronto, Canada"),
    "Peugeot Design Center": (48.7729096, 2.2207472, "medium", "Centre technique Stellantis de Vélizy, Vélizy-Villacoublay, France"),
    "Phoenix Art Museum": (33.4667737, -112.0740656, "high", "Phoenix Art Museum, Arizona, United States"),
    "Federal Building": (37.7812810, -122.3974525, "high", "San Francisco Federal Building, California, United States"),
    "Zug Bahnhof": (47.1737633, 8.5167031, "high", "Zug railway station, Switzerland"),
    "VNG AG": (51.3617455, 12.4298435, "high", "VNG – Verbundnetz Gas Aktiengesellschaft, Braunstraße, Leipzig, Germany"),
    "Aria / CityCenter": (36.1074973, -115.1767724, "high", "ARIA Resort & Casino, Las Vegas, Nevada, United States"),
    "Indianapolis Museum of Art (Newfields)": (39.8183271, -86.1867615, "high", "Indianapolis Museum of Art at Newfields, Indiana, United States"),
    "Lenbachhaus": (48.1462165, 11.5580516, "high", "Lenbachhaus, Munich, Germany"),
    "Magasin III": (59.3428478, 18.1201735, "high", "Magasin 3, Frihamnsgatan, Stockholm, Sweden"),
    "LT Blouin Foundation": (51.5101256, -0.2177329, "high", "Louise Blouin Foundation, Olaf Street, London, United Kingdom"),
    "Principal Financial Group": (41.5868654, -93.6249494, "high", "Principal Financial Group, Des Moines, Iowa, United States"),
    "Arts Bibliotech": (41.3875806, 2.1816694, "high", "Deuce Coop, Carrer Comerç 23, Barcelona, Spain"),
    "Museum of Contemporary Art": (34.0532882, -118.2506228, "high", "Museum of Contemporary Art, South Grand Avenue, Los Angeles, California, United States"),
    "Franklin Park Conservatory": (39.9686381, -82.9509815, "high", "Franklin Park Conservatory and Botanical Gardens, Ohio, United States"),
    "Amanzoe (Aman Resorts)": (37.3653027, 23.1349532, "high", "Amanzoe, Greece"),
    "Bonte Museum": (33.2450797, 126.4128504, "high", "Bonte Museum, Jeju, South Korea"),
    "Private (Baker Residence)": (41.0262654, -73.6284630, "medium", "Greenwich, Connecticut, United States"),
    "DFA Office": (30.2978093, -97.8303419, "medium", "6300 Bee Cave Road, Austin, Texas, United States"),
}


def normalize(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def fetch_json(url):
    retries = [0, 2, 5]
    error = None
    for delay in retries:
        if delay:
            time.sleep(delay)
        try:
            req = urllib.request.Request(url, headers=HEADERS)
            with urllib.request.urlopen(req, timeout=30) as response:
                return json.load(response)
        except HTTPError as exc:
            error = exc
    if error:
        raise error
    raise RuntimeError(f"failed to fetch {url}")


def photon_search(query):
    url = "https://photon.komoot.io/api/?limit=3&q=" + urllib.parse.quote(query)
    data = fetch_json(url)
    results = []
    for feature in data.get("features", []):
        props = feature.get("properties", {})
        lng, lat = feature.get("geometry", {}).get("coordinates", [None, None])
        results.append(
            {
                "lat": lat,
                "lng": lng,
                "label": ", ".join(
                    part for part in [
                        props.get("name"),
                        props.get("street"),
                        props.get("city"),
                        props.get("state"),
                        props.get("country"),
                    ] if part
                ),
                "kind": props.get("type") or props.get("osm_key") or "",
            }
        )
    return results


def resolve_coordinates(title, venue, city, country):
    if venue in MANUAL_COORDS:
        lat, lng, confidence, label = MANUAL_COORDS[venue]
        return lat, lng, confidence, label, "manual"

    query = ", ".join(part for part in [venue, city, country] if part)
    try:
        results = photon_search(query)
    except Exception:
        results = []

    if results:
        top = results[0]
        confidence = "medium"
        label_text = (top["label"] or "").lower()
        if venue.lower() in label_text:
            confidence = "high"
        elif city and city.split(",")[0].strip().lower() in label_text:
            confidence = "low"
        return top["lat"], top["lng"], confidence, top["label"], "photon"

    if city and country:
        try:
            results = photon_search(f"{city}, {country}")
        except Exception:
            results = []
        if results:
            top = results[0]
            return top["lat"], top["lng"], "low", top["label"], "photon_city"

    return None, None, "unknown", "", "none"


def category_from_type(type_name):
    t = normalize(type_name).lower()
    if "architecture" in t:
        return "architecture"
    if "ganzfeld" in t or "installation" in t or "projection" in t or "perceptual cell" in t:
        return "installation"
    if "skyspace" in t or "land art" in t or "wedgework" in t or "tunnel" in t:
        return "site-specific"
    return "permanent"


def access_from_text(access_text):
    text = normalize(access_text).lower()
    if "private" in text or "limited" in text or "unknown" in text:
        return "limited"
    if "appointment" in text or "reservation" in text or "ticket" in text or "guests" in text:
        return "reservation"
    return "public"


def main():
    wb = load_workbook(INPUT_XLSX, data_only=True)
    ws = wb["Site-Specific Works"]

    items = []
    venue_cache = {}

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        title, year, type_name, venue, city, country, public_access, source, status = row
        venue = normalize(venue)
        city = normalize(city)
        country = normalize(country)
        cache_key = (venue, city, country)
        if cache_key not in venue_cache:
            venue_cache[cache_key] = resolve_coordinates(normalize(title), venue, city, country)
            time.sleep(0.15)

        lat, lng, confidence, geocode_label, geocode_source = venue_cache[cache_key]

        items.append(
            {
                "id": idx,
                "name": normalize(title),
                "nameJa": normalize(title),
                "year": normalize(year),
                "location": ", ".join(part for part in [venue, city, country] if part),
                "locationJa": ", ".join(part for part in [venue, city, country] if part),
                "country": country,
                "category": category_from_type(type_name),
                "type": normalize(type_name),
                "venue": venue,
                "cityRegion": city,
                "access": access_from_text(public_access),
                "accessText": normalize(public_access),
                "status": normalize(status),
                "lat": lat,
                "lng": lng,
                "confidence": confidence,
                "description": f"{normalize(type_name)} at {venue}. Status: {normalize(status)}",
                "source": normalize(source),
                "geocodeLabel": geocode_label,
                "geocodeSource": geocode_source,
            }
        )

    OUTPUT_JSON.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {len(items)} works to {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
